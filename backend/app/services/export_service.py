#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
数据导出服务
支持导出订单、车辆、节点、路线等数据为Excel格式
"""

import io
import logging
from datetime import datetime
from typing import Dict, List, Optional
from app.models import db
from app.models import Order, Vehicle, Node, Route

logger = logging.getLogger(__name__)

# 尝试导入 openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExportService:
    """数据导出服务"""
    
    def __init__(self):
        if not HAS_OPENPYXL:
            logger.warning("openpyxl 未安装，Excel导出功能不可用")
    
    def export_orders(self, order_ids: List[int] = None, filters: Dict = None) -> bytes:
        """
        导出订单数据
        
        Args:
            order_ids: 指定订单ID列表
            filters: 筛选条件
        
        Returns:
            Excel文件字节流
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl 未安装")
        
        # 查询订单
        query = Order.query
        
        if order_ids:
            query = query.filter(Order.id.in_(order_ids))
        
        if filters:
            if filters.get('status'):
                query = query.filter(Order.status == filters['status'])
            if filters.get('start_date'):
                query = query.filter(Order.created_at >= filters['start_date'])
            if filters.get('end_date'):
                query = query.filter(Order.created_at <= filters['end_date'])
        
        orders = query.order_by(Order.created_at.desc()).all()
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '订单数据'
        
        # 表头
        headers = ['订单ID', '订单号', '客户名称', '客户电话', '货物名称', 
                   '货物类型', '重量(kg)', '体积(m³)', '优先级', '状态',
                   '取货节点', '送货节点', '分配车辆', '创建时间', '备注']
        
        self._write_header(ws, headers)
        
        # 数据行
        for row_idx, order in enumerate(orders, start=2):
            ws.cell(row=row_idx, column=1, value=order.id)
            ws.cell(row=row_idx, column=2, value=order.order_number)
            ws.cell(row=row_idx, column=3, value=order.customer_name)
            ws.cell(row=row_idx, column=4, value=order.customer_phone)
            ws.cell(row=row_idx, column=5, value=order.cargo_name)
            ws.cell(row=row_idx, column=6, value=order.cargo_type)
            ws.cell(row=row_idx, column=7, value=order.weight)
            ws.cell(row=row_idx, column=8, value=order.volume)
            ws.cell(row=row_idx, column=9, value=order.priority)
            ws.cell(row=row_idx, column=10, value=order.status)
            ws.cell(row=row_idx, column=11, value=order.pickup_node.name if order.pickup_node else '')
            ws.cell(row=row_idx, column=12, value=order.delivery_node.name if order.delivery_node else '')
            ws.cell(row=row_idx, column=13, value=order.vehicle.plate_number if order.vehicle else '')
            ws.cell(row=row_idx, column=14, value=order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else '')
            ws.cell(row=row_idx, column=15, value=order.notes or '')
        
        self._auto_width(ws)
        
        return self._to_bytes(wb)
    
    def export_vehicles(self, vehicle_ids: List[int] = None) -> bytes:
        """导出车辆数据"""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl 未安装")
        
        query = Vehicle.query
        if vehicle_ids:
            query = query.filter(Vehicle.id.in_(vehicle_ids))
        
        vehicles = query.order_by(Vehicle.created_at.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = '车辆数据'
        
        headers = ['车辆ID', '车牌号', '车辆类型', '品牌', '型号',
                   '载重(吨)', '容积(m³)', '司机姓名', '司机电话', '状态', '备注']
        
        self._write_header(ws, headers)
        
        for row_idx, v in enumerate(vehicles, start=2):
            ws.cell(row=row_idx, column=1, value=v.id)
            ws.cell(row=row_idx, column=2, value=v.plate_number)
            ws.cell(row=row_idx, column=3, value=v.vehicle_type)
            ws.cell(row=row_idx, column=4, value=v.brand)
            ws.cell(row=row_idx, column=5, value=v.model)
            ws.cell(row=row_idx, column=6, value=v.load_capacity)
            ws.cell(row=row_idx, column=7, value=v.volume_capacity)
            ws.cell(row=row_idx, column=8, value=v.driver_name)
            ws.cell(row=row_idx, column=9, value=v.driver_phone)
            ws.cell(row=row_idx, column=10, value=v.status)
            ws.cell(row=row_idx, column=11, value=v.notes or '')
        
        self._auto_width(ws)
        
        return self._to_bytes(wb)
    
    def export_nodes(self, node_ids: List[int] = None) -> bytes:
        """导出节点数据"""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl 未安装")
        
        query = Node.query
        if node_ids:
            query = query.filter(Node.id.in_(node_ids))
        
        nodes = query.order_by(Node.created_at.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = '节点数据'
        
        headers = ['节点ID', '节点名称', '节点类型', '地址', '经度', '纬度',
                   '联系人', '联系电话', '容量', '状态', '备注']
        
        self._write_header(ws, headers)
        
        for row_idx, n in enumerate(nodes, start=2):
            ws.cell(row=row_idx, column=1, value=n.id)
            ws.cell(row=row_idx, column=2, value=n.name)
            ws.cell(row=row_idx, column=3, value=n.type)
            ws.cell(row=row_idx, column=4, value=n.address)
            ws.cell(row=row_idx, column=5, value=n.longitude)
            ws.cell(row=row_idx, column=6, value=n.latitude)
            ws.cell(row=row_idx, column=7, value=n.contact_name)
            ws.cell(row=row_idx, column=8, value=n.contact_phone)
            ws.cell(row=row_idx, column=9, value=n.capacity)
            ws.cell(row=row_idx, column=10, value=n.status)
            ws.cell(row=row_idx, column=11, value=n.notes or '')
        
        self._auto_width(ws)
        
        return self._to_bytes(wb)
    
    def export_routes(self, route_ids: List[int] = None) -> bytes:
        """导出路线数据"""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl 未安装")
        
        query = Route.query
        if route_ids:
            query = query.filter(Route.id.in_(route_ids))
        
        routes = query.order_by(Route.created_at.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = '路线数据'
        
        headers = ['路线ID', '路线名称', '起点', '终点', '距离(km)', 
                   '时长(min)', '成本(元)', '状态', '备注']
        
        self._write_header(ws, headers)
        
        for row_idx, r in enumerate(routes, start=2):
            ws.cell(row=row_idx, column=1, value=r.id)
            ws.cell(row=row_idx, column=2, value=r.name)
            ws.cell(row=row_idx, column=3, value=r.start_node.name if r.start_node else '')
            ws.cell(row=row_idx, column=4, value=r.end_node.name if r.end_node else '')
            ws.cell(row=row_idx, column=5, value=r.distance)
            ws.cell(row=row_idx, column=6, value=r.duration)
            ws.cell(row=row_idx, column=7, value=r.total_cost)
            ws.cell(row=row_idx, column=8, value=r.status)
            ws.cell(row=row_idx, column=9, value=r.notes or '')
        
        self._auto_width(ws)
        
        return self._to_bytes(wb)
    
    def export_all(self) -> Dict[str, bytes]:
        """导出所有数据"""
        return {
            'orders.xlsx': self.export_orders(),
            'vehicles.xlsx': self.export_vehicles(),
            'nodes.xlsx': self.export_nodes(),
            'routes.xlsx': self.export_routes()
        }
    
    def export_data(self, data_type: str = 'nodes', format: str = 'xlsx') -> tuple:
        """
        通用导出数据方法
        
        Args:
            data_type: 数据类型 (nodes, routes, vehicles, orders)
            format: 导出格式 (xlsx, csv)
        
        Returns:
            (file_data, filename) 元组
        """
        export_map = {
            'nodes': (self.export_nodes, '节点数据'),
            'routes': (self.export_routes, '路线数据'),
            'vehicles': (self.export_vehicles, '车辆数据'),
            'orders': (self.export_orders, '订单数据')
        }
        
        if data_type not in export_map:
            raise ValueError(f'不支持的导出类型: {data_type}')
        
        export_func, name = export_map[data_type]
        file_data = export_func()
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{name}_{timestamp}.xlsx'
        
        # 如果是 CSV 格式，需要转换
        if format == 'csv':
            file_data = self._xlsx_to_csv(file_data)
            filename = filename.replace('.xlsx', '.csv')
        
        return (io.BytesIO(file_data), filename)
    
    def export_transport_report(self, start_date: str = None, end_date: str = None, format: str = 'xlsx') -> tuple:
        """
        导出运输报表
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            format: 导出格式
        
        Returns:
            (file_data, filename) 元组
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl 未安装")
        
        # 查询订单
        query = Order.query
        if start_date:
            query = query.filter(Order.created_at >= start_date)
        if end_date:
            query = query.filter(Order.created_at <= end_date)
        
        orders = query.order_by(Order.created_at.desc()).all()
        
        # 创建报表
        wb = Workbook()
        ws = wb.active
        ws.title = '运输报表'
        
        # 标题
        ws['A1'] = '物流运输报表'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A3'] = f'统计周期: {start_date or "全部"} ~ {end_date or "至今"}'
        
        # 统计数据
        total_orders = len(orders)
        completed = len([o for o in orders if o.status == 'delivered'])
        total_revenue = sum(o.actual_cost or o.estimated_cost or 0 for o in orders)
        
        ws['A5'] = '统计概览'
        ws['A5'].font = Font(size=12, bold=True)
        ws['A6'] = f'总订单数: {total_orders}'
        ws['A7'] = f'已完成: {completed}'
        ws['A8'] = f'总收入: ¥{total_revenue:.2f}'
        
        # 明细表
        ws['A10'] = '订单明细'
        ws['A10'].font = Font(size=12, bold=True)
        
        headers = ['订单号', '客户', '货物', '取货点', '送货点', '状态', '金额', '创建时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        for row_idx, o in enumerate(orders, 12):
            ws.cell(row=row_idx, column=1, value=o.order_number)
            ws.cell(row=row_idx, column=2, value=o.customer_name)
            ws.cell(row=row_idx, column=3, value=o.cargo_name)
            ws.cell(row=row_idx, column=4, value=o.pickup_node.name if o.pickup_node else '')
            ws.cell(row=row_idx, column=5, value=o.delivery_node.name if o.delivery_node else '')
            ws.cell(row=row_idx, column=6, value=o.status)
            ws.cell(row=row_idx, column=7, value=o.actual_cost or o.estimated_cost or 0)
            ws.cell(row=row_idx, column=8, value=o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else '')
        
        self._auto_width(ws)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'运输报表_{timestamp}.xlsx'
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return (buffer, filename)
    
    def generate_template(self, template_type: str) -> tuple:
        """
        生成导入模板
        
        Args:
            template_type: 模板类型 (nodes, orders, vehicles, routes)
        
        Returns:
            (file_data, filename) 元组
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl 未安装")
        
        wb = Workbook()
        ws = wb.active
        
        template_map = {
            'nodes': {
                'title': '节点导入模板',
                'headers': ['节点名称*', '节点类型', '地址', '经度', '纬度', 
                           '联系人', '联系电话', '容量', '备注']
            },
            'orders': {
                'title': '订单导入模板',
                'headers': ['订单号', '客户名称*', '客户电话', '货物名称', '货物类型',
                           '重量(kg)', '体积(m³)', '优先级', '状态', '取货节点', 
                           '送货节点', '预计成本', '备注']
            },
            'vehicles': {
                'title': '车辆导入模板',
                'headers': ['车牌号*', '车辆类型', '品牌', '型号', '载重(吨)',
                           '容积(m³)', '司机姓名', '司机电话', '状态', '备注']
            },
            'routes': {
                'title': '路线导入模板',
                'headers': ['路线名称*', '起点节点', '终点节点', '距离(km)', 
                           '时长(分钟)', '成本', '备注']
            }
        }
        
        if template_type not in template_map:
            raise ValueError(f'不支持的模板类型: {template_type}')
        
        config = template_map[template_type]
        ws.title = config['title']
        
        # 写入标题
        ws['A1'] = config['title']
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A1:{get_column_letter(len(config["headers"]))}1')
        
        # 写入说明
        ws['A2'] = '说明: 带 * 号的为必填字段'
        ws['A2'].font = Font(color='FF0000', italic=True)
        
        # 写入表头
        for col, header in enumerate(config['headers'], 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 添加示例行
        example_row = 4
        if template_type == 'nodes':
            ws.cell(row=example_row, column=1, value='北京仓库')
            ws.cell(row=example_row, column=2, value='warehouse')
            ws.cell(row=example_row, column=3, value='北京市朝阳区xxx')
        elif template_type == 'orders':
            ws.cell(row=example_row, column=2, value='张三')
            ws.cell(row=example_row, column=4, value='电子产品')
            ws.cell(row=example_row, column=5, value='general')
        elif template_type == 'vehicles':
            ws.cell(row=example_row, column=1, value='京A12345')
            ws.cell(row=example_row, column=2, value='货车')
        elif template_type == 'routes':
            ws.cell(row=example_row, column=1, value='北京-上海')
            ws.cell(row=example_row, column=2, value='北京仓库')
            ws.cell(row=example_row, column=3, value='上海仓库')
        
        self._auto_width(ws)
        
        filename = f'{config["title"]}.xlsx'
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return (buffer, filename)
    
    def backup_database(self) -> tuple:
        """
        备份数据库
        
        Returns:
            (file_data, filename) 元组
        """
        import shutil
        
        # 获取数据库路径
        db_path = db.engine.url.database
        if not db_path:
            raise RuntimeError('无法获取数据库路径')
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'database_backup_{timestamp}.db'
        
        # 复制数据库文件
        buffer = io.BytesIO()
        shutil.copy2(db_path, buffer.name if hasattr(buffer, 'name') else 'temp_backup.db')
        
        with open(db_path, 'rb') as f:
            buffer = io.BytesIO(f.read())
        
        buffer.seek(0)
        return (buffer, filename)
    
    def _xlsx_to_csv(self, xlsx_data: bytes) -> bytes:
        """将 xlsx 转换为 csv"""
        import csv
        
        wb = self._load_workbook(io.BytesIO(xlsx_data))
        ws = wb.active
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
        
        return output.getvalue().encode('utf-8-sig')
    
    def _load_workbook(self, file_data):
        """加载工作簿"""
        from openpyxl import load_workbook
        return load_workbook(file_data)
    
    def _write_header(self, ws, headers: List[str]):
        """写入表头"""
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _auto_width(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _to_bytes(self, wb) -> bytes:
        """将工作簿转换为字节流"""
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


# 单例实例
_export_service = None


def get_export_service() -> ExportService:
    """获取导出服务实例"""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service
