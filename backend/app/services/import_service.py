#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
数据导入服务
支持从Excel导入订单、车辆、节点等数据
"""

import io
import logging
from datetime import datetime
from typing import Dict, List, Tuple
from app.models import db
from app.models import Order, Vehicle, Node, Route

logger = logging.getLogger(__name__)

# 尝试导入 openpyxl
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ImportService:
    """数据导入服务"""
    
    def __init__(self):
        if not HAS_OPENPYXL:
            logger.warning("openpyxl 未安装，Excel导入功能不可用")
    
    def import_orders(self, file_data: bytes) -> Dict:
        """
        导入订单数据
        
        Args:
            file_data: Excel文件字节流
        
        Returns:
            导入结果 {'success': bool, 'imported': int, 'failed': int, 'errors': list}
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl 未安装")
        
        wb = load_workbook(io.BytesIO(file_data))
        ws = wb.active
        
        result = {
            'success': True,
            'imported': 0,
            'failed': 0,
            'errors': []
        }
        
        # 读取数据（跳过表头）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # 解析行数据
                order_data = self._parse_order_row(row)
                
                # 验证必填字段
                if not order_data.get('customer_name'):
                    result['errors'].append(f'第{row_idx}行: 缺少客户名称')
                    result['failed'] += 1
                    continue
                
                # 查找节点
                pickup_node = None
                delivery_node = None
                
                if order_data.get('pickup_node_name'):
                    pickup_node = Node.query.filter_by(name=order_data['pickup_node_name']).first()
                if order_data.get('delivery_node_name'):
                    delivery_node = Node.query.filter_by(name=order_data['delivery_node_name']).first()
                
                # 生成订单号
                order_number = f'ORD{datetime.now().strftime("%Y%m%d%H%M%S")}{row_idx}'
                
                # 创建订单
                order = Order(
                    order_number=order_number,
                    customer_name=order_data['customer_name'],
                    customer_phone=order_data.get('customer_phone', ''),
                    pickup_node_id=pickup_node.id if pickup_node else None,
                    delivery_node_id=delivery_node.id if delivery_node else None,
                    cargo_name=order_data.get('cargo_name'),
                    cargo_type=order_data.get('cargo_type', 'general'),
                    weight=order_data.get('weight', 0),
                    volume=order_data.get('volume', 0),
                    priority=order_data.get('priority', 'normal'),
                    status='pending',
                    notes=order_data.get('notes')
                )
                
                db.session.add(order)
                result['imported'] += 1
                
            except Exception as e:
                result['errors'].append(f'第{row_idx}行: {str(e)}')
                result['failed'] += 1
        
        db.session.commit()
        
        return result
    
    def import_vehicles(self, file_data: bytes) -> Dict:
        """导入车辆数据"""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl 未安装")
        
        wb = load_workbook(io.BytesIO(file_data))
        ws = wb.active
        
        result = {
            'success': True,
            'imported': 0,
            'failed': 0,
            'errors': []
        }
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                vehicle_data = self._parse_vehicle_row(row)
                
                if not vehicle_data.get('plate_number'):
                    result['errors'].append(f'第{row_idx}行: 缺少车牌号')
                    result['failed'] += 1
                    continue
                
                # 检查是否已存在
                if Vehicle.query.filter_by(plate_number=vehicle_data['plate_number']).first():
                    result['errors'].append(f'第{row_idx}行: 车牌号已存在')
                    result['failed'] += 1
                    continue
                
                vehicle = Vehicle(
                    plate_number=vehicle_data['plate_number'],
                    vehicle_type=vehicle_data.get('vehicle_type'),
                    brand=vehicle_data.get('brand'),
                    model=vehicle_data.get('model'),
                    load_capacity=vehicle_data.get('load_capacity'),
                    volume_capacity=vehicle_data.get('volume_capacity'),
                    driver_name=vehicle_data.get('driver_name'),
                    driver_phone=vehicle_data.get('driver_phone'),
                    notes=vehicle_data.get('notes')
                )
                
                db.session.add(vehicle)
                result['imported'] += 1
                
            except Exception as e:
                result['errors'].append(f'第{row_idx}行: {str(e)}')
                result['failed'] += 1
        
        db.session.commit()
        
        return result
    
    def import_nodes(self, file_data: bytes) -> Dict:
        """导入节点数据"""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl 未安装")
        
        wb = load_workbook(io.BytesIO(file_data))
        ws = wb.active
        
        result = {
            'success': True,
            'imported': 0,
            'failed': 0,
            'errors': []
        }
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                node_data = self._parse_node_row(row)
                
                if not node_data.get('name'):
                    result['errors'].append(f'第{row_idx}行: 缺少节点名称')
                    result['failed'] += 1
                    continue
                
                node = Node(
                    name=node_data['name'],
                    type=node_data.get('type', 'customer'),
                    address=node_data.get('address'),
                    longitude=node_data.get('longitude'),
                    latitude=node_data.get('latitude'),
                    contact_name=node_data.get('contact_name'),
                    contact_phone=node_data.get('contact_phone'),
                    capacity=node_data.get('capacity'),
                    notes=node_data.get('notes')
                )
                
                db.session.add(node)
                result['imported'] += 1
                
            except Exception as e:
                result['errors'].append(f'第{row_idx}行: {str(e)}')
                result['failed'] += 1
        
        db.session.commit()
        
        return result
    
    def _parse_order_row(self, row: tuple) -> Dict:
        """解析订单行"""
        return {
            'customer_name': row[2] if len(row) > 2 else None,
            'customer_phone': row[3] if len(row) > 3 else None,
            'cargo_name': row[4] if len(row) > 4 else None,
            'cargo_type': row[5] if len(row) > 5 else None,
            'weight': row[6] if len(row) > 6 else None,
            'volume': row[7] if len(row) > 7 else None,
            'priority': row[8] if len(row) > 8 else None,
            'pickup_node_name': row[10] if len(row) > 10 else None,
            'delivery_node_name': row[11] if len(row) > 11 else None,
            'notes': row[14] if len(row) > 14 else None
        }
    
    def _parse_vehicle_row(self, row: tuple) -> Dict:
        """解析车辆行"""
        return {
            'plate_number': row[1] if len(row) > 1 else None,
            'vehicle_type': row[2] if len(row) > 2 else None,
            'brand': row[3] if len(row) > 3 else None,
            'model': row[4] if len(row) > 4 else None,
            'load_capacity': row[5] if len(row) > 5 else None,
            'volume_capacity': row[6] if len(row) > 6 else None,
            'driver_name': row[7] if len(row) > 7 else None,
            'driver_phone': row[8] if len(row) > 8 else None,
            'notes': row[10] if len(row) > 10 else None
        }
    
    def _parse_node_row(self, row: tuple) -> Dict:
        """解析节点行"""
        return {
            'name': row[1] if len(row) > 1 else None,
            'type': row[2] if len(row) > 2 else None,
            'address': row[3] if len(row) > 3 else None,
            'longitude': row[4] if len(row) > 4 else None,
            'latitude': row[5] if len(row) > 5 else None,
            'contact_name': row[6] if len(row) > 6 else None,
            'contact_phone': row[7] if len(row) > 7 else None,
            'capacity': row[8] if len(row) > 8 else None,
            'notes': row[10] if len(row) > 10 else None
        }

# 单例实例
_import_service = None


def get_import_service() -> ImportService:
    """获取导入服务实例"""
    global _import_service
    if _import_service is None:
        _import_service = ImportService()
    return _import_service
