#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
数据分析 API 路由
"""

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.services.analytics_service import get_analytics_service
import logging
import io

logger = logging.getLogger(__name__)

analytics_bp = Blueprint('analytics', __name__)


@analytics_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def get_dashboard():
    """获取仪表盘数据"""
    try:
        service = get_analytics_service()
        result = service.get_dashboard_metrics()
        return jsonify(result)
    except Exception as e:
        logger.error(f"获取仪表盘数据失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/trend', methods=['GET'])
@jwt_required()
def get_trend():
    """
    获取趋势分析
    
    Query params:
        start_date: 开始日期 (YYYY-MM-DD)
        end_date: 结束日期 (YYYY-MM-DD)
        granularity: 粒度 (daily, weekly, monthly)
    """
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        granularity = request.args.get('granularity', 'daily')
        
        service = get_analytics_service()
        result = service.get_trend_analysis(start_date, end_date, granularity)
        return jsonify(result)
    except Exception as e:
        logger.error(f"获取趋势分析失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/cost', methods=['GET'])
@jwt_required()
def get_cost_analysis():
    """
    获取成本分析
    
    Query params:
        period: 分析周期 (week, month, quarter, year)
    """
    try:
        period = request.args.get('period', 'month')
        
        service = get_analytics_service()
        result = service.get_cost_analysis(period)
        return jsonify(result)
    except Exception as e:
        logger.error(f"获取成本分析失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/routes', methods=['GET'])
@jwt_required()
def get_route_performance():
    """获取路线性能分析"""
    try:
        service = get_analytics_service()
        result = service.get_route_performance()
        return jsonify(result)
    except Exception as e:
        logger.error(f"获取路线性能分析失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/vehicles', methods=['GET'])
@jwt_required()
def get_vehicle_performance():
    """获取车辆性能分析"""
    try:
        service = get_analytics_service()
        result = service.get_vehicle_performance()
        return jsonify(result)
    except Exception as e:
        logger.error(f"获取车辆性能分析失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/report', methods=['GET'])
@jwt_required()
def generate_report():
    """
    生成运营报表
    
    Query params:
        type: 报表类型 (daily, weekly, monthly)
        start_date: 开始日期
        end_date: 结束日期
    """
    try:
        report_type = request.args.get('type', 'daily')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        service = get_analytics_service()
        result = service.generate_report(report_type, start_date, end_date)
        return jsonify(result)
    except Exception as e:
        logger.error(f"生成报表失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/predict', methods=['GET'])
@jwt_required()
def predict_demand():
    """
    预测需求
    
    Query params:
        days: 预测天数 (默认7天)
    """
    try:
        days = request.args.get('days', 7, type=int)
        days = min(max(days, 1), 30)  # 限制1-30天
        
        service = get_analytics_service()
        result = service.predict_demand(days)
        return jsonify(result)
    except Exception as e:
        logger.error(f"预测需求失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/export/excel', methods=['GET'])
@jwt_required()
def export_excel():
    """
    导出 Excel 报表
    
    Query params:
        type: 报表类型 (daily, weekly, monthly)
        start_date: 开始日期
        end_date: 结束日期
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        report_type = request.args.get('type', 'daily')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        service = get_analytics_service()
        report = service.generate_report(report_type, start_date, end_date)
        
        if not report.get('success'):
            return jsonify(report), 400
        
        # 创建 Excel 工作簿
        wb = Workbook()
        
        # 概览页
        ws_summary = wb.active
        ws_summary.title = '运营概览'
        
        # 标题样式
        title_font = Font(size=16, bold=True)
        header_font = Font(size=12, bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # 报表标题
        ws_summary['A1'] = f'物流系统运营报表 - {report_type}'
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:D1')
        
        # 生成时间
        ws_summary['A2'] = f"生成时间: {report['report_info']['generated_at']}"
        
        # 关键指标
        metrics = report.get('summary', {})
        ws_summary['A4'] = '关键指标'
        ws_summary['A4'].font = header_font
        
        metric_names = {
            'total_orders': '总订单数',
            'completed_orders': '已完成订单',
            'pending_orders': '待处理订单',
            'total_revenue': '总收入(元)',
            'total_cost': '总成本(元)',
            'profit_margin': '利润率(%)',
            'vehicle_utilization': '车辆利用率(%)',
            'today_orders': '今日订单',
            'week_orders': '本周订单',
            'month_orders': '本月订单'
        }
        
        row = 5
        for key, label in metric_names.items():
            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = metrics.get(key, 0)
            row += 1
        
        # 趋势数据页
        ws_trend = wb.create_sheet('趋势数据')
        trend_data = report.get('trend_analysis', {})
        
        ws_trend['A1'] = '趋势分析'
        ws_trend['A1'].font = header_font
        
        ws_trend['A3'] = '总订单'
        ws_trend['B3'] = trend_data.get('total_orders', 0)
        ws_trend['A4'] = '总收入'
        ws_trend['B4'] = trend_data.get('total_revenue', 0)
        ws_trend['A5'] = '总成本'
        ws_trend['B5'] = trend_data.get('total_cost', 0)
        ws_trend['A6'] = '总利润'
        ws_trend['B6'] = trend_data.get('total_profit', 0)
        
        # 成本分析页
        ws_cost = wb.create_sheet('成本分析')
        cost_data = report.get('cost_analysis', {})
        
        ws_cost['A1'] = '成本分析'
        ws_cost['A1'].font = header_font
        
        ws_cost['A3'] = '总成本'
        ws_cost['B3'] = cost_data.get('total_cost', 0)
        ws_cost['A4'] = '单均成本'
        ws_cost['B4'] = cost_data.get('avg_cost_per_order', 0)
        
        breakdown = cost_data.get('cost_breakdown', {})
        ws_cost['A6'] = '成本构成'
        ws_cost['A7'] = '燃油成本'
        ws_cost['B7'] = breakdown.get('fuel', 0)
        ws_cost['A8'] = '过路费'
        ws_cost['B8'] = breakdown.get('toll', 0)
        ws_cost['A9'] = '人工成本'
        ws_cost['B9'] = breakdown.get('labor', 0)
        
        # 优化建议页
        ws_reco = wb.create_sheet('优化建议')
        recommendations = report.get('recommendations', [])
        
        ws_reco['A1'] = '优化建议'
        ws_reco['A1'].font = header_font
        
        for i, rec in enumerate(recommendations, start=3):
            ws_reco[f'A{i}'] = rec
        
        # 调整列宽
        for ws in [ws_summary, ws_trend, ws_cost, ws_reco]:
            for col in range(1, 5):
                ws.column_dimensions[get_column_letter(col)].width = 20
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 返回文件
        filename = f"logistics_report_{report_type}_{start_date or 'latest'}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except ImportError:
        return jsonify({
            'success': False, 
            'error': 'Excel 导出需要安装 openpyxl: pip install openpyxl'
        }), 500
    except Exception as e:
        logger.error(f"导出 Excel 失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/export/pdf', methods=['GET'])
@jwt_required()
def export_pdf():
    """
    导出 PDF 报表
    
    Query params:
        type: 报表类型 (daily, weekly, monthly)
        start_date: 开始日期
        end_date: 结束日期
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        
        report_type = request.args.get('type', 'daily')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        service = get_analytics_service()
        report = service.generate_report(report_type, start_date, end_date)
        
        if not report.get('success'):
            return jsonify(report), 400
        
        # 创建 PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, 
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18)
        heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=14)
        
        elements = []
        
        # 标题
        elements.append(Paragraph(f'物流系统运营报表 - {report_type}', title_style))
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(f"生成时间: {report['report_info']['generated_at']}", styles['Normal']))
        elements.append(Spacer(1, 1*cm))
        
        # 关键指标表格
        elements.append(Paragraph('关键指标', heading_style))
        elements.append(Spacer(1, 0.3*cm))
        
        metrics = report.get('summary', {})
        metric_names = {
            '总订单数': metrics.get('total_orders', 0),
            '已完成订单': metrics.get('completed_orders', 0),
            '待处理订单': metrics.get('pending_orders', 0),
            '总收入(元)': metrics.get('total_revenue', 0),
            '总成本(元)': metrics.get('total_cost', 0),
            '利润率(%)': metrics.get('profit_margin', 0)
        }
        
        table_data = [['指标', '数值']]
        for key, value in metric_names.items():
            table_data.append([key, str(value)])
        
        table = Table(table_data, colWidths=[8*cm, 6*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 1*cm))
        
        # 优化建议
        elements.append(Paragraph('优化建议', heading_style))
        elements.append(Spacer(1, 0.3*cm))
        
        recommendations = report.get('recommendations', [])
        for rec in recommendations:
            elements.append(Paragraph(f'• {rec}', styles['Normal']))
        
        # 生成 PDF
        doc.build(elements)
        output.seek(0)
        
        filename = f"logistics_report_{report_type}_{start_date or 'latest'}.pdf"
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    except ImportError:
        return jsonify({
            'success': False,
            'error': 'PDF 导出需要安装 reportlab: pip install reportlab'
        }), 500
    except Exception as e:
        logger.error(f"导出 PDF 失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500